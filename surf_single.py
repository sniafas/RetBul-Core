#!/usr/bin/env python

import cv2
import argparse,sys
import numpy as np
import math
from openpyxl import Workbook
import pandas as pd

def wdata(kp1,d1,kp2,d2,inliers,acc):

	wb = Workbook()
	ws = wb.active
	ws.title = "Data Collection"
	ws.append(["Query Image","Test Image","X1","Y1","X2","Y2","Scale1","Orientation1","Scale2","Orientation2","Inliers","Score"])
	kpid1 = 2
	kpid2 = 2 
	ws.cell(row=kpid1, column=1, value=img1Path) #house_id
	ws.cell(row=kpid2, column=2, value=img2Path) #image_id
	for k1 in kp1:
		
		ws.cell(row=kpid1, column=3, value=k1.pt[0]) #x1
		ws.cell(row=kpid1, column=4, value=k1.pt[1]) #y1			
		ws.cell(row=kpid1, column=7, value=k1.size) #scale1
		ws.cell(row=kpid1, column=8, value=k1.angle) #orientation1
		kpid1 = kpid1 + 1

	for k2 in kp2:
		
		ws.cell(row=kpid2, column=5, value=k2.pt[0]) #x2
		ws.cell(row=kpid2, column=6, value=k2.pt[1]) #y2			
		ws.cell(row=kpid2, column=9, value=k2.size) #scale2
		ws.cell(row=kpid2, column=10, value=k2.angle) #orientation2
		kpid2 = kpid2 + 1

	ws.cell(row=2, column=11, value=inliers) #score
	ws.cell(row=2, column=12, value=acc) #score
	wb.save(filename = 'surf_single_output.xls')
	pd.read_excel('surf_single_output.xls', sheetname='Data Collection').to_csv('surf_single_output.csv', index=False)

def drawKeypoint(img, p):
	
	for i in range(0,len(p)):
		
		x = int(round(p[i].pt[0]))
		y = int(round(p[i].pt[1]))
		center = (x,y)
			
		radius = round(p[i].size/2) # KeyPoint::size is a diameter
		
		
		#draw the circles around keypoints with the keypoints size
		cv2.circle( img, center, int(radius), (0,0,100), 1)
		
		#draw orientation of the keypoint, if it is applicable		
		if p[i].angle != -1 :
			
			srcAngleRad = p[i].angle * 3.14159/180;
			orient1 = int(round(math.cos(srcAngleRad)*radius ))
			orient2 = int(round(math.sin(srcAngleRad)*radius ))
			cv2.line( img, center, (x+orient1,y+orient2), (0,150,0), 1);

	return img

def filter_rawMatches(kp1, kp2, matches, ratio = 0.75):

	mkp1, mkp2 = [], []
	
	for r in range(len(matches)-1):

		if matches[r].distance < ratio * matches[r+1].distance:
			m = matches[r]
			mkp1.append(kp1[m.queryIdx])
			mkp2.append(kp2[m.trainIdx])
	
	p1 = np.float32([kp.pt for kp in mkp1])
	p2 = np.float32([kp.pt for kp in mkp2])
	kp_pairs = zip(mkp1, mkp2)	

	return p1,p2,kp_pairs		
		

if __name__ == '__main__':

	parser = argparse.ArgumentParser()

	parser.add_argument('-img1', action='append', dest='im1',             
	                    help='Query Image')
	parser.add_argument('-img2', action='append', dest='im2',
	                    help='Test Image')
	parser.add_argument('-H', action='store', dest='hss', type=int,
	                     default=200,help='Hessian Matrix <0-5000>')
	parser.add_argument('-nO', action='store', dest='nO', type=int,
	                    default=3,help='Number of Octaves <1-4>')
	parser.add_argument('-nL', action='store', dest='nL', type=int,
	                     default=3,help='Number of Octave Layers <3-6>')
	parser.add_argument('-u', action='store_true',default=False,
	                    dest='upright',help='SURF Upright')
	parser.add_argument('-e', action='store_true',default=False,
	                    dest='e',help='64/128 bit Length')
	parser.add_argument('-kpfixed', action='store_true',default=False,
	                    dest='kpfixed',help='Fixed Keypoint Colour')
	parser.add_argument('-v', action='store_true',default=False,
	                    dest='v',help='Save image to file')
	parser.add_argument('-o', action='store_true',default=False,
	                    dest='o',help='Save data to CSV')
	results = parser.parse_args()


	if results.im1:
		img1Path = str(results.im1)[2:-2]
	else:
		parser.print_help()
		print "-img1: Query Image"	
		sys.exit(1)

	if results.im2:	# single version
		img2Path = str(results.im2)[2:-2]
	else:
		parser.print_help()
		print "-img2: Test Image"	
		sys.exit(1)	

	print("\n================")
	print("Hessian", results.hss)
	print("Octaves", results.nO)
	print("Layers", results.nL)
	print("================")
	print("Processing...\n")

	## SURF features and descriptor
	surf = cv2.xfeatures2d.SURF_create(results.hss,results.nO,results.nL,results.e,results.upright)
	## #----------------- # ##
	## Read, Resize, Grayscale Query Image ##
	img1 = cv2.resize(cv2.imread(img1Path, 1), (480, 640))
	img1Gray = cv2.cvtColor(img1,cv2.COLOR_RGB2GRAY)	
	kp1, d1 = surf.detectAndCompute(img1Gray, None)
	
	## #----------------- # ##
	## Read, Resize, Grayscale Test Image ## 
	img2 = cv2.resize(cv2.imread(img2Path, 1), (480, 640))
	img2Gray = cv2.cvtColor(img2,cv2.COLOR_RGB2GRAY)	
	kp2, d2 = surf.detectAndCompute(img2Gray, None)


	## # Use BFMatcher, Euclidian distance, Eliminate Multiples # ##
	bf = cv2.BFMatcher(cv2.NORM_L2,crossCheck=True)
	raw_matches = bf.match(d1,d2)
	src_points, dst_points, kp_pairs = filter_rawMatches(kp1,kp2,raw_matches)

	print 'Matching tentative points in image1: %d, image2: %d' % (len(src_points), len(dst_points))
	
	## # ----------------# ##
	## # Homography # ##

	print 'Homography\n'

	if len(kp_pairs) > 4:
		
		Homography, status = cv2.findHomography(src_points, dst_points, cv2.RANSAC, 5.0)
		
		inliers = np.count_nonzero(status)
		percent = float(inliers) / len(kp_pairs)
	
		print "# Inliers %d out of %d tentative pairs" % (inliers,len(kp_pairs))
		print '{percent:.2%}'.format(percent= percent)

	else:
		print "Not enough correspondenses"

	#Output CSV
	if results.o:
		wdata(kp1,d1,kp2,d2,inliers,percent)

	## Verbose Results
	if results.v: 		
		
		img1kp = img1
		img2kp = img2
		if results.kpfixed:
			img1kp = drawKeypoint(img1kp,kp1)
			img2kp = drawKeypoint(img2kp,kp2)
		else:			
			cv2.drawKeypoints(img1kp,kp1,img1kp,None,flags=cv2.DRAW_MATCHES_FLAGS_DRAW_RICH_KEYPOINTS)
			cv2.drawKeypoints(img2kp,kp2,img2kp,None,flags=cv2.DRAW_MATCHES_FLAGS_DRAW_RICH_KEYPOINTS)
			print("Default Keypoints")	
		
		cv2.imshow('Query',img1kp)
		cv2.imshow('Test',img2kp)
		imgTentMatches = cv2.drawMatches(img1kp,kp1,img2kp,kp2,raw_matches,None, flags=2)
		cv2.imshow('Tentative Matches',imgTentMatches)

		try:
			h1, w1, z1 = img1.shape[:3]
			h2, w2, z2 = img2.shape[:3]
			img3 = np.zeros((max(h1, h2), w1+w2,z1), np.uint8)
			img3[:h1, :w1, :z1] = cv2.resize(cv2.imread(img1Path, 1), (480, 640))
			img3[:h2, w1:w1+w2, :z2] = cv2.resize(cv2.imread(img2Path, 1), (480, 640))
			
			p1 = np.int32([kpp[0].pt for kpp in kp_pairs])
			p2 = np.int32([kpp[1].pt for kpp in kp_pairs]) + (w1, 0)
			
			for (x1, y1), (x2, y2), inlier in zip(p1,p2, status):
				if inlier:
					cv2.circle(img3, (x1, y1), 2, (0,250,0), 5)
					cv2.circle(img3, (x2, y2), 2, (0,250,0), 5)
					cv2.line(img3, (x1, y1), (x2, y2), (255,100,0),2)
				else:
					cv2.line(img3, (x1-2, y1-2), (x1+2, y1+2), (0, 0, 255), 3)
					cv2.line(img3, (x1-2, y1+2), (x1+2, y1-2), (0, 0, 255), 3)
					cv2.line(img3, (x2-2, y2-2), (x2+2, y2+2), (0, 0, 255), 3)
					cv2.line(img3, (x2-2, y2+2), (x2+2, y2-2), (0, 0, 255), 3)
		
		except (RuntimeError, TypeError, NameError):
			print "Not enough Inliers"

		cv2.imshow('SURF Match + Inliers',img3)
		cv2.waitKey(0)
		cv2.destroyAllWindows()