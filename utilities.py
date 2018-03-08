#!/usr/bin/env python

import cv2
import pandas as pd
import math
from openpyxl import Workbook, load_workbook


class Utilities:

	def __init__(self):
		
		self.initWrite()

	def initWrite(self):
		
		self.head = 1
		self.wb = Workbook()
		self.ws = self.wb.active			
		self.ws.title = "Dataset"

	def closeWrite(self,outFolder,image,desc):

		self.fname = outFolder + str(desc) + '_' + str(image[:-4]) + ".xls"
		self.wb.save(self.fname)
		pd.read_excel(self.fname, sheetname="Dataset").to_csv(outFolder + desc + '_' + image[8:-4] + '_test.csv', index=False)

	def writeFile(self,keypoints,descriptors,qImage,tImage,inliers,percent,size):

		self.ws.cell(row=self.head, column=1, value=qImage)
		self.ws.cell(row=self.head, column=2, value=tImage)
		self.ws.cell(row=self.head, column=7, value=inliers) 
		self.ws.cell(row=self.head, column=8, value=percent)
		self.ws.cell(row=self.head, column=9, value=size)
		for k in keypoints:			
			self.ws.cell(row=self.head, column=3, value=k.pt[0]) #x2
			self.ws.cell(row=self.head, column=4, value=k.pt[1]) #y2			
			self.ws.cell(row=self.head, column=5, value=k.size) #scale2
			self.ws.cell(row=self.head, column=6, value=k.angle) #orientation2
			self.head = self.head + 1

	def drawKeypoint(self, img, p):
		
		for i in range(0,len(p)):
			
			x = int(round(p[i].pt[0]))
			y = int(round(p[i].pt[1]))
			center = (x,y)				
			radius = round(p[i].size/2) # KeyPoint::size is a diameter			
			
			#draw the circles around keypoints with the keypoints size
			cv2.circle( img, center, int(radius), (0,0,100), 1)

			#draw orientation of the keypoint, if it is applicable		
			if p[i].angle != -1 :				
				srcAngleRad = p[i].angle * 3.14159/180;
				orient1 = int(round(math.cos(srcAngleRad)*radius ))
				orient2 = int(round(math.sin(srcAngleRad)*radius ))
				cv2.line( img, center, (x+orient1,y+orient2), (0,150,0), 1)
		return img